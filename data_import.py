import re
from openpyxl import *


def data_from_isc_seminar(filename, ) -> []:
    data = []

    wb = load_workbook(filename=filename)
    sheet = wb['Worksheet']
    # sheet = wb['Junior']

    # Create a dictionary of column names
    columns = {}
    index = 0
    for COL in sheet.iter_cols(1, sheet.max_column):
        columns[COL[0].value] = index
        index += 1

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append({
            "name_last": str(row[columns["Nachname"]]).strip(),
            "name_first": str(row[columns["Vorname"]]).strip(),
            "birth_date":  row[columns["Geb.Dat"]].strftime("%d.%m.%Y") if row[columns["Geb.Dat"]] is not None else "",
            "street": str(row[columns["Strasse"]]).strip() if row[columns["Strasse"]] is not None else "",
            "postcode": str(row[columns["Plz"]]).strip() if row[columns["Plz"]] is not None else "",
            "city": str(row[columns["Ort"]]).strip() if row[columns["Ort"]] is not None else "",
            "country": "DEU",
            "mail": str(row[columns["E-Mail"]]).strip().lower(),
            "rank": str(re.findall("Junior|Bronze|Silber|Gold", row[columns["Gewünschtes Abzeichen"]])[0]).strip(),
            "repetition": "",
            "first_registration": ""
        })

    return data


def data_from_group_register(filename) -> []:
    data = []

    ranks = ["Junior", "Bronze", "Silber", "Gold", "DSTA"]

    wb = load_workbook(filename=filename)

    for rank in ranks:
        try:
            sheet = wb[rank]
        except Exception:
            continue

        # Create a dictionary of column names
        columns = {}
        index = 0
        for COL in sheet.iter_cols(1, sheet.max_column):
            columns[COL[0].value] = index
            index += 1

        for row in sheet.iter_rows(min_row=2, max_row=50, values_only=True):
            data.append({
                "name_last": str(row[columns["Nachname"]]).strip() if row[columns["Nachname"]] is not None else "",
                "name_first": str(row[columns["Vorname"]]).strip() if row[columns["Vorname"]] is not None else "",
                "birth_date": str(row[columns["GebDatum"]]).strip() if row[columns["GebDatum"]] is not None else "",
                # "birth_date": row[columns["GebDatum"]].strftime("%d.%m.%Y") if row[columns["GebDatum"]] is not None else "",
                "street": str(row[columns["Straße"]]).strip() if row[columns["Straße"]] is not None else "",
                "postcode": str(row[columns["PLZ"]]).strip() if row[columns["PLZ"]] is not None else "",
                "city": str(row[columns["Wohnort"]]).strip() if row[columns["Wohnort"]] is not None else "",
                "country": "DEU",
                "mail": str(row[columns["E-Mail"]]).strip().lower(),
                "rank": rank,
                "repetition": "",
                "first_registration": ""
            })
    return data
